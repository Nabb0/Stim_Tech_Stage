import re
import os
import openpyxl
import phonenumbers

#!Metodi Generali
#?Check list
def creaLista():
    words = ["Nome:", "Cognome:", "IBAN:", "Mittente:", "Destinatario:", "Richiedente:", "Firmatario:", "Testimone:", "Notaio:", "Debitore:", "Creditore:", "Procuratore:", "Beneficiario:", "Garante:", "Contraente:", "Testatore:", "Beneficiario effettivo:", "Amministratore:", "Titolare:", "Proprietario:", "Inquilino:", "Reclamante:", "Responsabile legale:", "Agente:", "Giudice:", "Avvocato:", "Proponente:", "Curatore:", "Supplente:", "Padrone di casa:", "Testimone oculare:", "Assicurato:", "Beneficiario fiduciario:", "Beneficiario di polizza:", "Coobbligato:", "Curatore speciale:", "Datore di lavoro:", "Difensore:", "Donante:", "Esecutore testamentario:", "Garante finanziario:", "Garante solidale:", "Intermediario:", "Legatario:", "Mediatore:", "Perito:", "Premiante:", "Promittente:", "Proposto:", "Referente:", "Richiedente asilo:", "Soggetto obbligato:", "Testimone esperto:", "Tribunale:", "Ufficiale pubblico:", "Valutatore:", "Venditore:", "Acquirente:", "Conduttore:", "Lasciatario:", "Locatore:", "Concessionario:", "Contraente principale:", "Contraente secondario:", "Locatario:", "Mandante:", "Mandatario:", "Offerta:", "Ricevente:", "Sofferente:", "Subappaltatore:", "Subentrante:", "Sublocatario:", "Sublocatore:", "Testimone di nozze:", "Promittente venditore:", "Promissario acquirente:"]

    # Duplica le parole senza i due punti
    words_no_colon = [word.replace(":", "") for word in words]

    # Unisci le due liste
    words_combined = words + words_no_colon

    return words_combined

#?Trova la stringa passata nel testo
def find_string_in_text(strings, file_path):
    lAssociazioni = []
    with open(file_path, 'r') as file:
        text = file.readlines()

    for string in strings:
        for i, line in enumerate(text):
            if string in line:
                if i >= 2:  # Assicurati che ci siano almeno due linee precedenti
                    lAssociazioni.append([text[i-2].rstrip('\n'), text[i-1].rstrip('\n'), line.rstrip('\n')])
    
    for i in range(len(lAssociazioni)):
        add_to_excel(lAssociazioni[i])

#?Aggiungi a excel
def add_to_excel(list):
    # Creazione del nuovo file Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Intestazioni delle colonne
    column_headers = ["Nome", "Cognome", "IBAN"]
    sheet.append(column_headers)

    # Popolamento delle colonne con i dati
    for i, data in enumerate(list):
        info = data.split()
        nome = ""
        cognome = ""
        iban = ""

        # Estrazione delle informazioni dal testo
        for item in info:
            if item.startswith("Nome:"):
                nome = info[1]
                # Scrittura dei dati nelle colonne
                sheet.cell(row=i+2, column=1).value = nome.strip('\n')
            elif item.startswith("Cognome:"):
                cognome = info[1]
                sheet.cell(row=i+2, column=2).value = cognome.strip('\n')
            elif item.startswith("IBAN:"):
                iban = info[1]
                sheet.cell(row=i+2, column=3).value = iban.strip('\n')

    # Salvataggio del file Excel
    workbook.save("Excel/File/Output/output.xlsx")


#!Estrazione degli iban
def extract_iban(text):
    pattern = r"\b(?:IT|SM)[0-9]{2}[A-Za-z][0-9]{22}\b"  # Pattern to recognize an Italian or Sanmarinese IBAN

    ibans = re.findall(pattern, text)
    return ibans

def extract_iban_from_txt(text):
    ibans = extract_iban(text)
    return ibans

def visIban(directory):
    file_list = os.listdir(directory)
    for file in file_list:
        if file.endswith(".txt"):
            file_path = os.path.join(directory, file)
            with open(file_path, 'r') as file:
                text = file.read()
            ibans = extract_iban_from_txt(text)

            find_string_in_text(ibans, file_path)


#!Estrazione codice fiscale
def extract_codice_fiscale(text):
    pattern = r"\b[A-Z]{6}\d{2}[A-Z]\d{2}[A-Z]\d{3}[A-Z]\b"  # Pattern per riconoscere un codice fiscale italiano

    codici_fiscali = re.findall(pattern, text)
    return codici_fiscali

def extract_codice_fiscale_from_txt(text):
    codici_fiscali = extract_codice_fiscale(text)
    return codici_fiscali

def visCodFisc(directory):
    file_list = os.listdir(directory)
    for file in file_list:
        if file.endswith(".txt"):
            file_path = os.path.join(directory, file)
            with open(file_path, 'r') as file:
                text = file.read()
            codici_fiscali = extract_codice_fiscale_from_txt(text)


#!Estrazione Nome e Cognome
def extract_nomi(text):
    textSplit = text.split()
    j = 0
    lista = []
    controllo = creaLista()
    for i in textSplit:
        dato = ''
        for k in controllo: 
            if i == k:
                dato = textSplit[j + 1]
                lista.append(dato)
                if textSplit[j + 2][0].isupper() and textSplit[j + 2][1].islower() and textSplit[j + 2] != 'C.F:' and textSplit[j + 2] != 'Cognome:':
                    lista.append(textSplit[j + 2])
        j += 1
    return lista

def extract_nomi_from_txt(text):
    nomi = extract_nomi(text)
    return nomi

def visNome(directory):
    file_list = os.listdir(directory)
    for file in file_list:
        if file.endswith(".txt"):
            file_path = os.path.join(directory, file)
            with open(file_path, 'r') as file:
                text = file.read()
            nomi = extract_nomi_from_txt(text)


#!Estrazione Date
def extract_dates_from_text(text):
    pattern = r"\b\d{4}[-/]\d{2}[-/]\d{2}\b"  # Pattern to recognize dates in "dd/mm/yyyy" or "dd-mm-yyyy" format

    dates = re.findall(pattern, text)
    return dates

def extract_dates_from_txt(text):
    dates = extract_dates_from_text(text)
    return dates

def visDate(directory):
    file_list = os.listdir(directory)
    for file in file_list:
        if file.endswith(".txt"):
            file_path = os.path.join(directory, file)
            with open(file_path, 'r') as file:
                text = file.read()
            dates = extract_dates_from_txt(text)


#!Estrai numero di telefono
def extract_phone_numbers_from_text(text):
    phone_numbers = []

    # Trova tutti i possibili numeri di telefono nel testo
    for match in phonenumbers.PhoneNumberMatcher(text, "IT"):
        phone_number = phonenumbers.format_number(match.number, phonenumbers.PhoneNumberFormat.INTERNATIONAL)
        phone_numbers.append(phone_number)

    return phone_numbers

def extract_phone_numbers_from_txt(text):
    phone_numbers = extract_phone_numbers_from_text(text)
    return phone_numbers

def visPhoneNumbers(directory):
    file_list = os.listdir(directory)
    for file in file_list:
        if file.endswith(".txt"):
            file_path = os.path.join(directory, file)
            with open(file_path, 'r') as file:
                text = file.read()
            phone_numbers = extract_phone_numbers_from_txt(text)


def main():
    directory = "E:/informatica/Visual Studio Code Projects/py projects/Stim Script/Txt/File/"
    visIban(directory)
    
    
main()
